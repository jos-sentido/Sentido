from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

prospectos = [
    {
        "nombre": "Tapicería Automotriz Del Parque",
        "nicho": "Tapicería automotriz",
        "direccion": "Av. del Chamizal #570 A, Col. Medrano, 44810 Guadalajara, Jal.",
        "telefono": "+52 33 3401 4788",
        "email": "",
        "facebook": "facebook.com/p/Tapicería-Automotriz-Del-Parque-100063657126095",
        "horario": "Lun-Vie 9:00-19:00 / Sáb 9:00-16:00",
        "presencia_actual": "Solo Facebook + directorios",
        "tiene_web": "No",
        "potencial_ingresos": "Medio-Alto: ticket alto en restauración + clientes recurrentes (talleres, agencias)",
        "angulo_propuesta": "Galería de antes/después + cotizador rápido vía WhatsApp + sello de 'Servicio a domicilio'",
    },
    {
        "nombre": "LECCSA Laboratorio Dental",
        "nicho": "Laboratorio dental (B2B)",
        "direccion": "Lago Pátzcuaro #766-B, 2do Piso, Col. Italia Providencia, 44648 Guadalajara, Jal.",
        "telefono": "+52 33 3641 1488",
        "email": "",
        "facebook": "facebook.com/leccsadental",
        "horario": "",
        "presencia_actual": "Solo Facebook + directorios",
        "tiene_web": "No",
        "potencial_ingresos": "Alto: B2B con dentistas, márgenes altos, cero estacionalidad",
        "angulo_propuesta": "Portal B2B con catálogo de prótesis (zirconia, e.max, CAD/CAM) + envío de casos por WhatsApp + tabla de tiempos de entrega",
    },
    {
        "nombre": "Lo Soldo Todo",
        "nicho": "Soldadura y herrería industrial",
        "direccion": "Av. 8 de Julio #1395, Col. Morelos, 44910 Guadalajara, Jal.",
        "telefono": "33 1228 7580",
        "email": "soldatodo@outlook.com",
        "facebook": "facebook.com/LOSOLDOTODO",
        "horario": "",
        "presencia_actual": "Directorios (Guialis, Mexicoo) + Facebook",
        "tiene_web": "No",
        "potencial_ingresos": "Medio-Alto: trabajos industriales, estructuras, mantenimiento",
        "angulo_propuesta": "Sitio con portafolio fotográfico de trabajos + formulario de cotización rápida + zonas de servicio en GDL",
    },
    {
        "nombre": "Fábrica de Muebles Alatorre",
        "nicho": "Fabricante de muebles a medida",
        "direccion": "Loma Honda #16, Tonalá, Jal.",
        "telefono": "+52 33 1016 9600",
        "email": "daniel.alatorre.salas98@gmail.com",
        "facebook": "facebook.com/p/Fabrica-de-muebles-Alatorre-100063617003659 | IG: muebles.alatorre",
        "horario": "",
        "presencia_actual": "Webnode (template gratuito) + Facebook + Instagram",
        "tiene_web": "No (subdominio webnode.mx, no es web propia)",
        "potencial_ingresos": "Alto: ticket alto en muebles a medida, clientes residenciales y hoteleros",
        "angulo_propuesta": "Migrar de webnode a dominio propio + catálogo por categoría (salas, recámaras, cocinas) + cotizador",
    },
    {
        "nombre": "Forrajes Tonalá",
        "nicho": "Forrajería / alimento para ganado",
        "direccion": "Calle Colón, Col. Centro, Tonalá, Jal.",
        "telefono": "33 3683 0247",
        "email": "",
        "facebook": "",
        "horario": "",
        "presencia_actual": "Solo directorios (Guiamexican, Mexicoo)",
        "tiene_web": "No",
        "potencial_ingresos": "Medio: clientes recurrentes (ganaderos, granjas) con compras semanales",
        "angulo_propuesta": "Catálogo de marcas (API, Purina, El Pedregal) + pedidos por WhatsApp + entrega a rancho",
    },
    {
        "nombre": "Forrajes Cornejo",
        "nicho": "Forrajería / alimento para ganado",
        "direccion": "Calle Periférico, Tonalá, Jal., CP 45412",
        "telefono": "33 3607 2896",
        "email": "",
        "facebook": "",
        "horario": "",
        "presencia_actual": "Solo directorios (Guialis, Comercio Empresa)",
        "tiene_web": "No",
        "potencial_ingresos": "Medio: forrajería de paso sobre periférico, alto tráfico de clientes",
        "angulo_propuesta": "Sitio con ubicación GPS + lista de precios actualizable + reserva de pedido por WhatsApp",
    },
    {
        "nombre": "Bordados GDL",
        "nicho": "Bordados industriales / uniformes",
        "direccion": "Calle Guanajuato #1653, Col. Villaguerrero, Guadalajara, Jal.",
        "telefono": "+52 33 1943 7953",
        "email": "bordadosgdl1@gmail.com",
        "facebook": "facebook.com/bordadosgdl1",
        "horario": "Lun-Sáb 6:00-14:00",
        "presencia_actual": "Solo Facebook",
        "tiene_web": "No",
        "potencial_ingresos": "Medio-Alto: clientes B2B (escuelas, empresas, restaurantes) con pedidos recurrentes de uniformes",
        "angulo_propuesta": "Portal B2B con simulador de bordado en logo + tabla de tiempos por volumen + portafolio por industria",
    },
    {
        "nombre": "DieselPoint Guadalajara",
        "nicho": "Laboratorio diésel (inyección, bombas, turbos)",
        "direccion": "Calz. Jesús González Gallo #2720-A, Col. El Rosario, 44898 Guadalajara, Jal.",
        "telefono": "33 2442 6629",
        "email": "",
        "facebook": "facebook.com/dieselpointmx",
        "horario": "Lun-Vie 9:00-18:30 / Sáb 9:00-14:00",
        "presencia_actual": "AllBiz + Facebook (sin web propia)",
        "tiene_web": "No",
        "potencial_ingresos": "Alto: ticket alto en reparación de bombas/turbos, clientes flotilleros",
        "angulo_propuesta": "Sitio con catálogo de servicios por marca (Cummins, John Deere, Bosch) + agenda de cita + sección B2B flotillas",
    },
    {
        "nombre": "Fabricantes de Muebles Cárdenas",
        "nicho": "Fabricante de muebles",
        "direccion": "Av. Tonaltecas #84 (Plaza Galerías del Sol L-5) y Calle Juárez #246, Tonalá Centro, Tonalá, Jal.",
        "telefono": "+52 33 1918 1850",
        "email": "",
        "facebook": "facebook.com/p/Fabricantes-de-muebles-Cárdenas-100054197484971",
        "horario": "",
        "presencia_actual": "Solo Facebook",
        "tiene_web": "No",
        "potencial_ingresos": "Medio-Alto: 2 sucursales en Tonalá, polo nacional de muebles, ticket alto",
        "angulo_propuesta": "Catálogo en línea con categorías (salas, comedores, recámaras) + ubicación de sucursales + envío a todo el país",
    },
    {
        "nombre": "KEYTRONIC Cerrajería Automotriz",
        "nicho": "Cerrajería automotriz especializada",
        "direccion": "Diagonal Manuel Cambre #1785 Local 23, Guadalajara, Jal.",
        "telefono": "+52 33 3870 0961",
        "email": "",
        "facebook": "facebook.com/keytronicCerrajeriaAutomotriz",
        "horario": "",
        "presencia_actual": "Solo Facebook",
        "tiene_web": "No",
        "potencial_ingresos": "Medio-Alto: programación de llaves con chip y mandos a distancia (ticket $2-8k MXN), urgencias 24/7",
        "angulo_propuesta": "Sitio con buscador 'mi auto / mi modelo / mi año' + tabla de precios por marca + botón emergencia 24h",
    },
]

wb = Workbook()
ws = wb.active
ws.title = "Prospectos GDL"

headers = [
    "#",
    "Nombre del negocio",
    "Nicho / Giro",
    "Dirección",
    "Teléfono",
    "Email",
    "Facebook / Redes",
    "Horario",
    "Presencia actual",
    "¿Tiene sitio web?",
    "Potencial de ingresos",
    "Ángulo de propuesta de diseño",
]

header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
cell_font = Font(name="Calibri", size=10)
alt_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
border = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
wrap_align = Alignment(wrap_text=True, vertical="top", horizontal="left")
center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align
    cell.border = border

for row_idx, p in enumerate(prospectos, start=2):
    values = [
        row_idx - 1,
        p["nombre"],
        p["nicho"],
        p["direccion"],
        p["telefono"],
        p["email"],
        p["facebook"],
        p["horario"],
        p["presencia_actual"],
        p["tiene_web"],
        p["potencial_ingresos"],
        p["angulo_propuesta"],
    ]
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = cell_font
        cell.alignment = wrap_align
        cell.border = border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

widths = [5, 30, 30, 50, 22, 32, 45, 28, 30, 18, 50, 60]
for col_idx, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 28
for r in range(2, len(prospectos) + 2):
    ws.row_dimensions[r].height = 90

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(prospectos) + 1}"

ws2 = wb.create_sheet("Notas y método")
notas = [
    ["BASE DE PROSPECTOS - GUADALAJARA, JALISCO"],
    [""],
    ["Fecha de curación: 26 de abril de 2026"],
    ["Total de prospectos: 10"],
    [""],
    ["Criterios de selección:"],
    ["1. Ubicación: Zona Metropolitana de Guadalajara (Guadalajara, Zapopan, Tonalá, Tlaquepaque)"],
    ["2. Nichos B2B / industriales con flujo de ingresos constante (no negocios de hype/moda)"],
    ["3. Sin sitio web propio (o solo template gratuito tipo Webnode/Ueniweb)"],
    ["4. Datos de contacto verificables: dirección + teléfono + presencia activa en redes"],
    [""],
    ["Distribución por giro:"],
    ["- Talleres mecánicos especializados (diésel, tapicería, cerrajería automotriz): 3"],
    ["- Fabricantes (muebles, soldadura/herrería): 3"],
    ["- B2B con clientes recurrentes (lab dental, bordados industriales): 2"],
    ["- Distribución agroindustrial (forrajerías): 2"],
    [""],
    ["Importante:"],
    ["- Estos datos provienen de búsquedas en directorios públicos y redes sociales (abril 2026)."],
    ["- Verifica el teléfono con una llamada antes del primer acercamiento comercial."],
    ["- Todos los negocios son operativos al momento de la búsqueda; algunos pueden tener cambios de ubicación."],
    [""],
    ["Recomendación de prospección:"],
    ["1. Llamar primero (no WhatsApp en frío) - estos negocios contestan teléfono."],
    ["2. Pedir hablar con el dueño / encargado - en estos giros suele ser una sola persona."],
    ["3. Llevar 2-3 ejemplos visuales de webs de su mismo giro (no de otros nichos)."],
    ["4. Cotizar paquete con dominio + hosting + WhatsApp Business + Google Business Profile."],
]
for row_idx, row_vals in enumerate(notas, start=1):
    for col_idx, val in enumerate(row_vals, start=1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if row_idx == 1:
            cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
        elif val and val[0:1].isdigit() and ". " in val[:4]:
            cell.font = Font(name="Calibri", size=10)
        elif val.endswith(":") or "Criterios" in val or "Distribución" in val or "Importante" in val or "Recomendación" in val:
            cell.font = Font(name="Calibri", size=11, bold=True)
        else:
            cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
ws2.column_dimensions["A"].width = 110

output_path = "/home/user/Sentido/prospectos_guadalajara.xlsx"
wb.save(output_path)
print(f"Archivo creado: {output_path}")
